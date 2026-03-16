"""Excel export service for billing reports."""
from datetime import datetime
from typing import List, Optional
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.database.models.utilizado import Utilizado, Factura
from app.database.models.entrada import Entrada


class ExcelExportService:
    """Service for exporting billing data to Excel."""

    @staticmethod
    def export_factura(factura: Factura) -> bytes:
        """Export single invoice to Excel.

        Args:
            factura: Factura instance to export.

        Returns:
            bytes: Excel file content.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Factura"

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

        # Title
        ws['A1'] = f"Factura N° {factura.numero_factura}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')

        # Invoice info
        ws['A3'] = "Fecha de Emisión:"
        ws['B3'] = factura.fecha_emision.strftime('%d/%m/%Y') if factura.fecha_emision else 'N/A'
        ws['A4'] = "Tipo:"
        ws['B4'] = factura.tipo_factura.value if hasattr(factura, 'tipo_factura') and factura.tipo_factura else 'B'
        ws['A5'] = "Estado:"
        ws['B5'] = factura.estado

        # Client info
        ws['A7'] = "Cliente:"
        ws['B7'] = factura.cliente.nombre if factura.cliente else 'N/A'
        ws['A8'] = "Dirección:"
        ws['B8'] = factura.cliente.direccion if factura.cliente and factura.cliente.direccion else 'N/A'

        # Headers for items
        headers = ["Ensayo", "Entrada", "Cantidad", "Precio Unit.", "Importe"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Get utilizados
        utilizados = Utilizado.query.filter_by(factura_id=factura.id).all()

        # Items
        row = 11
        for u in utilizados:
            ws.cell(row=row, column=1, value=u.ensayo.nombre_corto if u.ensayo else 'N/A')
            ws.cell(row=row, column=2, value=u.entrada.codigo if u.entrada else 'N/A')
            ws.cell(row=row, column=3, value=float(u.cantidad) if u.cantidad else 0)
            ws.cell(row=row, column=4, value=float(u.precio_unitario) if u.precio_unitario else 0)
            ws.cell(row=row, column=5, value=float(u.importe) if u.importe else 0)
            row += 1

        # Total
        ws.cell(row=row + 1, column=4, value="TOTAL:")
        ws.cell(row=row + 1, column=4).font = Font(bold=True)
        ws.cell(row=row + 1, column=5, value=float(factura.total) if factura.total else 0)
        ws.cell(row=row + 1, column=5).font = Font(bold=True)

        # Auto-adjust columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_reporte_cliente(
        cliente_id: int,
        fecha_inicio: Optional[datetime] = None,
        fecha_fin: Optional[datetime] = None
    ) -> bytes:
        """Export usage report for a client.

        Args:
            cliente_id: Client ID.
            fecha_inicio: Start date filter.
            fecha_fin: End date filter.

        Returns:
            bytes: Excel file content.
        """
        # Query utilizados for client
        query = Utilizado.query.join(
            Entrada, Utilizado.entrada_id == Entrada.id
        ).filter(
            Entrada.cliente_id == cliente_id
        )

        if fecha_inicio:
            query = query.filter(Utilizado.fecha_uso >= fecha_inicio)
        if fecha_fin:
            query = query.filter(Utilizado.fecha_uso <= fecha_fin)

        utilizados = query.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Header
        ws['A1'] = "Reporte de Uso - Cliente"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        headers = ["Fecha", "Ensayo", "Entrada", "Cantidad", "Precio", "Importe", "Estado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

        # Data
        for row_num, u in enumerate(utilizados, 4):
            ws.cell(row=row_num, column=1, value=u.fecha_uso.strftime('%d/%m/%Y') if u.fecha_uso else 'N/A')
            ws.cell(row=row_num, column=2, value=u.ensayo.nombre_corto if u.ensayo else 'N/A')
            ws.cell(row=row_num, column=3, value=u.entrada.codigo if u.entrada else 'N/A')
            ws.cell(row=row_num, column=4, value=float(u.cantidad) if u.cantidad else 0)
            ws.cell(row=row_num, column=5, value=float(u.precio_unitario) if u.precio_unitario else 0)
            ws.cell(row=row_num, column=6, value=float(u.importe) if u.importe else 0)
            ws.cell(row=row_num, column=7, value=u.estado)

        # Totals
        total_row = len(utilizados) + 5
        ws.cell(row=total_row, column=5, value="TOTAL:")
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=sum(float(u.importe) for u in utilizados))
        ws.cell(row=total_row, column=6).font = Font(bold=True)

        # Auto-adjust columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
